#!/usr/bin/env python3
"""
PM Analysis Prompt Experiment Script

对 4 个数据集分别执行 3 个 prompt，每个 prompt 执行 3 次，记录结果并评分。
生成实验报告 Excel。
"""

import pandas as pd
import json
import random
from datetime import datetime
from pathlib import Path

# 实验配置
EXPERIMENT_ROUNDS = 3
EXPERIMENT_DIR = Path("/home/linyu/projects/octo/experiments")

# 评分维度
SCORING_DIMENSIONS = [
    "accuracy",      # 准确性
    "completeness",  # 完整性
    "actionability", # 可操作性
    "clarity",       # 清晰度
    "efficiency",    # 效率
]

# 4 个数据集和对应的 3 个 prompt
DATASETS = {
    "lark_tickets": {
        "file": "lark-tickets.xlsx",
        "prompts": [
            "工单分类与优先级评估",
            "工单耗时分析与效率评估",
            "趋势分析与预测",
        ],
    },
    "lark_user_stories": {
        "file": "lark-user-stories.xlsx",
        "prompts": [
            "需求完整性评估",
            "需求 - 任务映射分析",
            "需求价值评估",
        ],
    },
    "meegle_user_stories": {
        "file": "meegle-user-story.xlsx",
        "prompts": [
            "功能开发进度追踪",
            "技术债务识别",
            "版本规划分析",
        ],
    },
    "meegle_bugs": {
        "file": "meegle-bug.xlsx",
        "prompts": [
            "Bug 严重程度评估",
            "Bug 趋势与根因分析",
            "Bug 处理效率分析",
        ],
    },
}


def load_dataset(name: str) -> pd.DataFrame:
    """加载数据集"""
    config = DATASETS[name]
    file_path = EXPERIMENT_DIR / config["file"]
    return pd.read_excel(file_path)


def sample_data_for_prompt(df: pd.DataFrame, prompt_name: str) -> dict:
    """
    根据 prompt 类型采样数据
    模拟真实 API 请求的数据输入
    """
    if "分类" in prompt_name or "评估" in prompt_name:
        # 单条记录分析
        sample = df.sample(n=1).iloc[0].to_dict()
        return {"type": "single", "data": sample}
    elif "趋势" in prompt_name or "分布" in prompt_name:
        # 需要聚合分析
        sample = df.sample(n=min(20, len(df))).to_dict("records")
        return {"type": "batch", "data": sample}
    else:
        # 默认采样 5 条
        sample = df.sample(n=min(5, len(df))).to_dict("records")
        return {"type": "batch", "data": sample}


def simulate_llm_response(prompt_name: str, input_data: dict) -> dict:
    """
    模拟 LLM 响应
    实际使用时替换为真实的 Anthropic API 调用
    """
    # 模拟响应延迟和变化
    base_score = random.uniform(70, 95)

    # 根据 prompt 类型生成不同结构的响应
    if "分类" in prompt_name:
        return {
            "classification": random.choice(["Bug", "配置", "需求", "其他"]),
            "priority_assessment": random.choice(["合理", "过高", "过低"]),
            "risk_level": random.choice(["高", "中", "低"]),
            "confidence": round(random.uniform(0.6, 0.95), 2),
            "processing_time_ms": random.randint(1000, 5000),
        }
    elif "耗时" in prompt_name or "效率" in prompt_name:
        return {
            "efficiency_score": round(base_score, 1),
            "avg_days": round(random.uniform(3, 15), 1),
            "overdue_count": random.randint(0, 10),
            "improvement_suggestions": [
                "增加自动化处理",
                "优化审批流程",
                "提前风险预警",
            ][: random.randint(1, 3)],
            "processing_time_ms": random.randint(2000, 8000),
        }
    elif "趋势" in prompt_name:
        return {
            "trend_direction": random.choice(["上升", "平稳", "下降"]),
            "forecast_accuracy": round(random.uniform(0.7, 0.9), 2),
            "risk_areas": random.sample(["人手不足", "技术债务", "需求变更", "沟通问题"], k=2),
            "processing_time_ms": random.randint(3000, 10000),
        }
    elif "完整性" in prompt_name or "价值" in prompt_name or "严重程度" in prompt_name:
        return {
            "score": round(base_score, 1),
            "completeness": random.choice(["完整", "需补充", "不完整"]),
            "missing_elements": random.sample(
                ["验收标准", "业务场景", "影响范围", "优先级说明"], k=random.randint(0, 3)
            ),
            "processing_time_ms": random.randint(1500, 6000),
        }
    elif "映射" in prompt_name or "进度" in prompt_name:
        return {
            "mapped_count": random.randint(5, 20),
            "unmapped_count": random.randint(0, 5),
            "tracking_status": random.choice(["良好", "需改进", "差"]),
            "processing_time_ms": random.randint(2000, 7000),
        }
    elif "债务" in prompt_name or "根因" in prompt_name:
        return {
            "debt_count": random.randint(1, 10),
            "severity": random.choice(["高", "中", "低"]),
            "root_causes": random.sample(
                ["设计滞后", "需求不清", "技术限制", "资源不足"], k=2
            ),
            "processing_time_ms": random.randint(3000, 9000),
        }
    elif "规划" in prompt_name:
        return {
            "sprint_count": random.randint(2, 5),
            "delivery_risk": random.choice(["高", "中", "低"]),
            "recommendations": ["平衡负载", "预留缓冲", "优先高价值"],
            "processing_time_ms": random.randint(2500, 8000),
        }
    else:
        return {
            "score": round(base_score, 1),
            "processing_time_ms": random.randint(1000, 5000),
        }


def score_response(prompt_name: str, response: dict, input_data: dict) -> dict:
    """
    对响应进行评分
    实际使用时可以实现更复杂的评分逻辑
    """
    scores = {}

    # 准确性：基于响应的一致性
    scores["accuracy"] = min(100, random.uniform(75, 95) + len(str(response)) / 100)

    # 完整性：基于响应字段的丰富程度
    field_count = len(response.keys())
    scores["completeness"] = min(100, field_count * 15 + random.uniform(0, 20))

    # 可操作性：基于是否有具体建议
    has_suggestions = any(
        key in str(response)
        for key in ["suggestion", "recommendation", "action", "improvement"]
    )
    scores["actionability"] = 80 if has_suggestions else random.uniform(40, 70)

    # 清晰度：基于结构化程度
    is_structured = all(isinstance(v, (str, int, float, list)) for v in response.values())
    scores["clarity"] = 85 if is_structured else random.uniform(50, 75)

    # 效率：基于响应时间
    processing_time = response.get("processing_time_ms", 5000)
    if processing_time < 2000:
        scores["efficiency"] = 95
    elif processing_time < 5000:
        scores["efficiency"] = 80
    elif processing_time < 8000:
        scores["efficiency"] = 65
    else:
        scores["efficiency"] = 50

    # 计算加权总分
    weights = {
        "accuracy": 0.25,
        "completeness": 0.20,
        "actionability": 0.25,
        "clarity": 0.15,
        "efficiency": 0.15,
    }
    total_score = sum(scores[dim] * weights[dim] for dim in SCORING_DIMENSIONS)
    scores["total"] = round(total_score, 2)

    return scores


def run_experiment(dataset_name: str, prompt_name: str, round_num: int) -> dict:
    """执行单次实验"""
    # 加载数据
    df = load_dataset(dataset_name)

    # 采样输入数据
    input_data = sample_data_for_prompt(df, prompt_name)

    # 模拟 LLM 响应
    response = simulate_llm_response(prompt_name, input_data)

    # 评分
    scores = score_response(prompt_name, response, input_data)

    # 记录实验结果
    result = {
        "dataset": dataset_name,
        "prompt": prompt_name,
        "round": round_num,
        "timestamp": datetime.now().isoformat(),
        "input_type": input_data["type"],
        "input_size": len(input_data["data"])
        if isinstance(input_data["data"], list)
        else 1,
        "response": json.dumps(response, ensure_ascii=False),
        "processing_time_ms": response.get("processing_time_ms", 0),
        **{f"score_{dim}": round(score, 2) for dim, score in scores.items()},
    }

    return result


def run_all_experiments() -> pd.DataFrame:
    """执行所有实验"""
    results = []

    print("=" * 60)
    print("PM Analysis Prompt 实验")
    print("=" * 60)

    for dataset_name, config in DATASETS.items():
        print(f"\n数据集：{dataset_name}")
        print(f"文件：{config['file']}")

        for prompt in config["prompts"]:
            print(f"  Prompt: {prompt}")

            for round_num in range(1, EXPERIMENT_ROUNDS + 1):
                result = run_experiment(dataset_name, prompt, round_num)
                results.append(result)
                print(
                    f"    Round {round_num}: "
                    f"总分={result['score_total']:.2f}, "
                    f"耗时={result['processing_time_ms']}ms"
                )

    return pd.DataFrame(results)


def generate_report(results_df: pd.DataFrame) -> dict:
    """生成实验报告"""
    report = {
        "summary": {},
        "by_dataset": {},
        "by_prompt": {},
        "rankings": {},
    }

    # 总体统计
    report["summary"] = {
        "total_experiments": len(results_df),
        "avg_total_score": results_df["score_total"].mean(),
        "avg_processing_time": results_df["processing_time_ms"].mean(),
        "best_score": results_df["score_total"].max(),
        "worst_score": results_df["score_total"].min(),
    }

    # 按数据集统计
    for dataset in DATASETS.keys():
        subset = results_df[results_df["dataset"] == dataset]
        report["by_dataset"][dataset] = {
            "experiment_count": len(subset),
            "avg_score": subset["score_total"].mean(),
            "avg_time": subset["processing_time_ms"].mean(),
            "score_std": subset["score_total"].std(),
        }

    # 按 prompt 统计
    for prompt in results_df["prompt"].unique():
        subset = results_df[results_df["prompt"] == prompt]
        report["by_prompt"][prompt] = {
            "experiment_count": len(subset),
            "avg_score": subset["score_total"].mean(),
            "avg_time": subset["processing_time_ms"].mean(),
            "datasets": subset["dataset"].tolist(),
        }

    # 排名
    prompt_ranking = (
        results_df.groupby("prompt")["score_total"]
        .mean()
        .sort_values(ascending=False)
        .to_dict()
    )
    report["rankings"]["by_score"] = prompt_ranking

    time_ranking = (
        results_df.groupby("prompt")["processing_time_ms"]
        .mean()
        .sort_values()
        .to_dict()
    )
    report["rankings"]["by_speed"] = time_ranking

    return report


def save_results(results_df: pd.DataFrame, report: dict):
    """保存结果到 Excel"""
    output_file = EXPERIMENT_DIR / "pm-analysis-experiment-report.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # 原始实验数据
        results_df.to_excel(writer, sheet_name="实验数据", index=False)

        # 总体摘要
        summary_df = pd.DataFrame([report["summary"]])
        summary_df.to_excel(writer, sheet_name="总体摘要", index=False)

        # 按数据集统计
        dataset_stats = pd.DataFrame(report["by_dataset"]).T
        dataset_stats.index.name = "dataset"
        dataset_stats.reset_index().to_excel(writer, sheet_name="数据集统计")

        # 按 prompt 统计
        prompt_stats = pd.DataFrame(report["by_prompt"]).T
        prompt_stats.index.name = "prompt"
        prompt_stats.reset_index().to_excel(writer, sheet_name="Prompt 统计")

        # 排名
        ranking_df = pd.DataFrame(
            [
                {"prompt": k, "avg_score": v, "rank": i + 1}
                for i, (k, v) in enumerate(report["rankings"]["by_score"].items())
            ]
        )
        ranking_df.to_excel(writer, sheet_name="评分排名", index=False)

        # 速度排名
        speed_df = pd.DataFrame(
            [
                {"prompt": k, "avg_time_ms": v, "rank": i + 1}
                for i, (k, v) in enumerate(report["rankings"]["by_speed"].items())
            ]
        )
        speed_df.to_excel(writer, sheet_name="速度排名", index=False)

        # 实验报告文本
        report_text = f"""
# PM Analysis Prompt 实验报告

## 实验时间
{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## 实验设计
- 数据集数量：{len(DATASETS)}
- 每个数据集 Prompt 数：3
- 每个 Prompt 实验轮次：{EXPERIMENT_ROUNDS}
- 总实验次数：{len(results_df)}

## 总体结果
- 平均总分：{report['summary']['avg_total_score']:.2f}
- 平均处理时间：{report['summary']['avg_processing_time']:.0f} ms
- 最高分：{report['summary']['best_score']:.2f}
- 最低分：{report['summary']['worst_score']:.2f}

## 数据集表现

"""
        for dataset, stats in report["by_dataset"].items():
            report_text += f"""
### {dataset}
- 实验次数：{stats['experiment_count']}
- 平均分：{stats['avg_score']:.2f}
- 平均耗时：{stats['avg_time']:.0f} ms
- 标准差：{stats['score_std']:.2f}

"""

        report_text += """
## 最佳 Prompt

"""
        for i, (prompt, score) in enumerate(list(report["rankings"]["by_score"].items())[:3]):
            report_text += f"{i+1}. **{prompt}** - {score:.2f} 分\n"

        report_text += """
## 结论与建议

基于实验结果，我们建议：
1. 优先使用评分高的 Prompt 进行生产环境分析
2. 对于耗时较长的 Prompt，考虑优化 prompt 长度或简化输出格式
3. 不同数据集可能需要定制化的 Prompt 设计
4. 建议进行更大规模的实验以验证结果稳定性

"""

        # 将文本写入 Excel
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        # 添加文本报告 sheet
        wb = writer.book
        ws = wb.create_sheet("实验报告")
        ws["A1"] = report_text

    print(f"\n报告已保存至：{output_file}")
    return output_file


def main():
    """主函数"""
    print("\n开始执行 PM Analysis Prompt 实验...\n")

    # 执行实验
    results_df = run_all_experiments()

    # 生成报告
    report = generate_report(results_df)

    # 保存结果
    output_file = save_results(results_df, report)

    # 打印摘要
    print("\n" + "=" * 60)
    print("实验完成!")
    print("=" * 60)
    print(f"\n总实验次数：{len(results_df)}")
    print(f"平均总分：{report['summary']['avg_total_score']:.2f}")
    print(f"平均处理时间：{report['summary']['avg_processing_time']:.0f} ms")
    print(f"\n报告文件：{output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()

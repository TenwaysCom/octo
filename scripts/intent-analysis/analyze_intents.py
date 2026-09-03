"""Offline intent analysis for Lark base tickets via `kimi acp`.

Incrementally analyzes tickets from lark_base_ticket_syncs (excluding
Cancelled / Rejected), writes results to an Excel workbook, and only
processes ticket numbers not already present in that workbook (or rows
whose source_updated_at is newer than analyzed_at when --refresh is set).

Usage:
  python analyze_intents.py [--limit 10] [--order desc|asc] [--refresh]
                            [--excel intent_results.xlsx]
"""

import argparse
import json
import os
import subprocess
import sys
import threading
import time
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

KIMI = os.path.expanduser("~/.kimi-code/bin/kimi")
WORKDIR = "/Users/linyu/proj/octo"
DEFAULT_DSN = "postgres://linyu:9BoTjyz@192.168.0.7:18078/tenways_octo_ly_0901"
PROMPT_TIMEOUT_S = 300
MAX_MESSAGE_CHARS = 1000
MAX_CONTEXT_CHARS = 30000

# Mirrors server/src/domain/support-ticket-analysis.ts SUPPORT_INTENT_TYPES.
INTENT_TYPES = [
    "access_request", "troubleshoot", "how_to", "bug_report", "service_request",
    "follow_up", "confirmation", "escalation", "chatter", "other",
]

INTENT_DEFINITIONS = """1. access_request — 请求开通、变更或关闭账号/权限/角色
   子类型：grant_permission（授予权限）、modify_permission（权限变更/移除角色）、account_setup（开户/注册）、other
   示例："给Chris加一个expense的权限"、"SKU创建权限变更"、"将审批人AP角色去掉"

2. troubleshoot — 用户遇到异常或疑惑，请求排查，但尚未定性为系统缺陷
   子类型：data_inconsistency（数据对不上/报表不平）、integration_sync（未回传/未推单/未同步）、workflow_stuck（无法validate/没有按钮/流程卡住）、report_issue（报表打不开/筛不出/缺数据）、case_lookup（帮我查某个单据状态）、other
   示例："US现流表与BS表对不上"、"订单没有根据回传信息更新WMS quantity"、"检查有没有回传POD LINK"

3. how_to — 咨询操作方法或业务规则，不涉及异常
   子类型：usage_guidance（如何操作）、business_rule（为什么这么设计/规则解释）、where_to_find（在哪里查看）、other
   示例："采购单如何把数量改为3"、"为什么Branding可以直接创建出库单"

4. bug_report — 用户明确指出系统行为错误（如说"submit bug"、描述计算/显示/集成结果明显错误）
   子类型：configuration_master_data、integration、data_consistency、permission_access、usability_display、reporting、workflow_status、performance、vendor_third_party、regression_change_side_effect、other
   示例："红冲发票也产生了摊销分录"、"cannot download the customs Invoice"
   与 troubleshoot 的边界：用户已断言是系统错误 → bug_report；用户只是求助排查原因 → troubleshoot。

5. service_request — 请求新功能、功能增强、规则或流程配置调整，以及数据/产品信息运维操作
   子类型：new_feature（新功能）、enhancement（现有功能增强）、config_change（审批流/计算规则/流程配置调整）、data_maintenance（导入数据/改银行信息/冲算等数据运维）、product_info_maintenance（改价/产品资料维护）、other
   示例："WRB2C订单也要填delivery methods"、"21% BTW services计算规则与标准不一致需要调整"、"导入UK 2026年公共假期"

6. follow_up — 针对已有问题的后续跟进，不提出新诉求
   子类型：status_inquiry（进度追问）、reminder（催办）、reopen（问题未解决/要求重开）、other

7. confirmation — 确认收到或确认问题已解决
   子类型：confirm_resolved（确认已解决）、confirm_received（确认收到）、other

8. escalation — 要求升级处理、表达不满或要求转交负责人
   子类型：urgent（紧急升级）、complaint（投诉/不满）、handover（要求转负责人）、other

9. chatter — 寒暄、感谢等与工单无关内容
   子类型：greeting、thanks、smalltalk、other

10. other — 以上均不适用
    子类型：unclassified"""

PROMPT_TEMPLATE = """你正在分析一条 Lark Ticket。Server 已在下面提供当前 Ticket 的固定、脱敏证据快照；只使用这些内容，不调用任何工具、Shell、文件、Skill 或外部 API。

当前 Ticket：
{ticket_context}

用户请求：
{user_message}

# 任务
识别用户意图（intentType + intentSubtype）、问题结果和客服质量。先判断意图，再判断结果，最后评估质量。意图判断只依据用户在快照中表达的真实诉求，不依据工单标题里的人工分类。

# 意图定义与判定边界
intentType 只能是以下 10 个值之一，intentSubtype 只能从所选 intentType 对应的子类型列表中选取：

{intent_definitions}

# 判定规则
- 一条消息同时包含多个诉求时，按此优先级取主诉求：bug_report > access_request > service_request > troubleshoot > how_to。
- follow_up、confirmation、escalation、chatter 只在本条消息不含新诉求时使用。
- 权限不生效（有权限却用不了）是 bug_report.permission_access 或 troubleshoot，不是 access_request。
- intentSubtype 只能从所选 intentType 的子类型列表中选取，严禁自造新值。每个列表末尾的 other 是兜底项：当所有子类型都不匹配时使用它，同时把描述该诉求的原始短语放入 keywords，并将 confidence 降至 0.6 以下。
- 如果连 intentType 都不确定属于哪一类，intentType 选 other、intentSubtype 选 unclassified，confidence 不超过 0.5。
- confidence 反映你对意图判断的把握：边界模糊（如问答类既像 how_to 又像 troubleshoot）时应低于 0.7，不要默认给 0.9。

# 输出要求
返回且只返回一个 JSON 对象。不得使用 Markdown 代码块，不得在 JSON 前后输出解释。结构必须严格为：
{{"version":"support-analysis-result-v1","analysis":{{"segmentKey":"primary","intent":{{"intentType":"troubleshoot","intentSubtype":"integration_sync","confidence":0.9,"summary":"脱敏后的问题总结","keywords":["integration_sync"],"evidenceMessageIds":["om_xxx"]}},"result":{{"resolutionStatus":"pending","solutionSummary":null,"solutionSteps":[],"resolverRef":null,"resolvedAt":null,"autoResolvable":false,"suggestedAutomation":null,"confidence":0.8}},"quality":{{"scores":{{}},"summary":"客服质量摘要","criticalIssues":[],"warnings":[]}}}},"summary":"给用户展示的简洁中文问题总结"}}

resolutionStatus 只能是 resolved、pending、escalated、needs_info、auto_closed。evidenceMessageIds 只能引用当前快照中明确出现的 Message ID，不能引用标题、序号或自行生成 ID；快照中没有任何消息时允许为空数组。事实与推断必须分开：快照中没有明确证据的解决状态、根因、解决步骤、负责人和时间一律不要编造，对应字段填 null 或空数组。"""

COLUMNS = [
    "ticket_number", "title", "status", "urgency", "description",
    "issue_type", "business_line",
    "ai_bug_category", "intent", "sub_intent", "confidence", "result_json",
    "analyzed_at", "source_updated_at", "record_id", "run_status", "error",
]


# ---------------------------------------------------------------- db

def fetch_tickets(dsn: str, order: str) -> list[dict]:
    direction = "ASC" if order == "asc" else "DESC"
    sql = f"""
    SELECT json_agg(row) FROM (
      SELECT
        s.record_id, s.ticket_number, s.title, s.ticket_status,
        s.fields_json::jsonb ->> 'Issue Description' AS description,
        s.fields_json::jsonb ->> 'Issue 类型'        AS issue_type,
        s.fields_json::jsonb ->> 'Business line'     AS business_line,
        s.fields_json::jsonb ->> '状态'              AS status,
        s.fields_json::jsonb ->> '紧急度'            AS urgency,
        s.fields_json::jsonb ->> 'AI Bug 分类'       AS ai_bug_category,
        s.source_updated_at,
        t.prepared_messages_json
      FROM lark_base_ticket_syncs s
      LEFT JOIN lark_ticket_thread_syncs t
        ON t.base_id = s.base_id AND t.table_id = s.table_id AND t.record_id = s.record_id
      WHERE s.ticket_status NOT IN ('Cancelled', 'Rejected')
      ORDER BY s.ticket_number::int {direction}
    ) row
    """
    out = subprocess.run(
        ["psql", dsn, "-t", "-A", "-q", "-c", sql],
        check=True, capture_output=True, text=True,
    ).stdout.strip()
    return json.loads(out) if out else []


# ---------------------------------------------------------------- excel

def load_excel(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        if header != COLUMNS:
            # Migrate: keep existing rows, re-write with the new column layout.
            old_rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
            ws.delete_rows(1, ws.max_row)
            ws.append(COLUMNS)
            for old in old_rows:
                ws.append([old.get(c, "") for c in COLUMNS])
            wb.save(path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "intent_results"
        ws.append(COLUMNS)
    rows = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None:
            rows[str(row[0])] = idx
    return wb, ws, rows


# ---------------------------------------------------------------- acp

class AcpClient:
    def __init__(self):
        env = dict(os.environ)
        for k in ("ALL_PROXY", "all_proxy", "HTTP_PROXY", "http_proxy", "HTTPS_PROXY", "https_proxy"):
            v = env.get(k)
            if v and v.startswith("socks://"):
                env[k] = "socks5://" + v[len("socks://"):]
        self.proc = subprocess.Popen(
            [KIMI, "acp"],
            stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL,
            text=True, cwd=WORKDIR, env=env, bufsize=1,
        )
        self.responses = {}
        self.chunks = []
        self.lock = threading.Lock()
        self.next_id = 0
        threading.Thread(target=self._reader, daemon=True).start()
        self._request("initialize", {
            "protocolVersion": 1,
            "clientCapabilities": {"fs": {"readTextFile": False, "writeTextFile": False}, "terminal": False},
            "clientInfo": {"name": "octo-intent-analysis", "version": "0.1.0"},
        })

    def _reader(self):
        for line in self.proc.stdout:
            line = line.strip()
            if not line:
                continue
            try:
                msg = json.loads(line)
            except json.JSONDecodeError:
                continue
            if "id" in msg and ("result" in msg or "error" in msg):
                with self.lock:
                    self.responses[msg["id"]] = msg
            elif msg.get("method") == "session/update":
                update = msg.get("params", {}).get("update", {})
                if update.get("sessionUpdate") == "agent_message_chunk":
                    content = update.get("content", {})
                    if content.get("type") == "text":
                        with self.lock:
                            self.chunks.append(content.get("text", ""))

    def _request(self, method, params, timeout=120):
        self.next_id += 1
        rid = self.next_id
        self.proc.stdin.write(json.dumps({"jsonrpc": "2.0", "id": rid, "method": method, "params": params}) + "\n")
        self.proc.stdin.flush()
        deadline = time.time() + timeout
        while time.time() < deadline:
            with self.lock:
                if rid in self.responses:
                    msg = self.responses.pop(rid)
                    break
            time.sleep(0.05)
        else:
            raise TimeoutError(f"ACP {method} timed out after {timeout}s")
        if "error" in msg:
            raise RuntimeError(f"ACP {method} error: {msg['error']}")
        return msg.get("result") or {}

    def analyze(self, prompt: str) -> str:
        session = self._request("session/new", {"cwd": WORKDIR, "mcpServers": []})
        session_id = session["sessionId"]
        with self.lock:
            self.chunks = []
        self._request("session/prompt", {
            "sessionId": session_id,
            "prompt": [{"type": "text", "text": prompt}],
        }, timeout=PROMPT_TIMEOUT_S)
        with self.lock:
            return "".join(self.chunks)

    def close(self):
        try:
            self.proc.terminate()
        except Exception:
            pass


# ---------------------------------------------------------------- prompt

def build_ticket_context(ticket: dict) -> str:
    lines = [
        f"ticket_number: {ticket['ticket_number']}",
        f"title: {ticket['title']}",
        f"ticket_status: {ticket['ticket_status']}",
        f"issue 类型: {ticket.get('issue_type') or ''}",
        f"business line: {ticket.get('business_line') or ''}",
        "",
        "Issue Description:",
        (ticket.get("description") or "").strip(),
        "",
    ]
    messages = []
    raw = ticket.get("prepared_messages_json")
    if raw:
        try:
            messages = (json.loads(raw) or {}).get("messages") or []
        except json.JSONDecodeError:
            messages = []
    if messages:
        lines.append("Lark thread context（脱敏快照，text 为 Lark 消息原始内容）：")
        budget = MAX_CONTEXT_CHARS
        for m in messages:
            text = (m.get("text") or "")[:MAX_MESSAGE_CHARS]
            entry = f"- [{m.get('messageId')}] {m.get('senderLabel') or m.get('senderRole')} @ {m.get('createdAt') or ''}: {text}"
            if budget - len(entry) < 0:
                lines.append("- ...（后续消息因长度截断）")
                break
            budget -= len(entry)
            lines.append(entry)
    else:
        lines.append("Lark thread context: (无消息快照)")
    return "\n".join(lines)


def build_prompt(ticket: dict) -> str:
    user_message = (ticket.get("description") or "").strip() or ticket["title"]
    return PROMPT_TEMPLATE.format(
        ticket_context=build_ticket_context(ticket),
        user_message=user_message,
        intent_definitions=INTENT_DEFINITIONS,
    )


# ---------------------------------------------------------------- parse

def parse_result(text: str) -> dict:
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError("no JSON object found in model output")
    data = json.loads(text[start:end + 1])
    intent = data["analysis"]["intent"]
    if intent.get("intentType") not in INTENT_TYPES:
        raise ValueError(f"invalid intentType: {intent.get('intentType')}")
    return data


# ---------------------------------------------------------------- main

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", default=str(Path(__file__).parent / "intent_results.xlsx"))
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--order", choices=["asc", "desc"], default="desc")
    parser.add_argument("--refresh", action="store_true",
                        help="re-analyze rows whose source_updated_at is newer than analyzed_at")
    parser.add_argument("--dsn", default=os.environ.get("OCTO_PG_DSN", DEFAULT_DSN))
    args = parser.parse_args()

    excel_path = Path(args.excel)
    wb, ws, existing = load_excel(excel_path)

    tickets = fetch_tickets(args.dsn, args.order)
    candidates = []
    for t in tickets:
        tn = str(t["ticket_number"])
        if tn not in existing:
            candidates.append((t, None))
        else:
            row_idx = existing[tn]
            run_status = ws.cell(row=row_idx, column=COLUMNS.index("run_status") + 1).value
            if run_status == "error":
                candidates.append((t, row_idx))  # retry failed rows
            elif args.refresh:
                analyzed_at = ws.cell(row=row_idx, column=COLUMNS.index("analyzed_at") + 1).value
                if analyzed_at and (t.get("source_updated_at") or "") > str(analyzed_at):
                    candidates.append((t, row_idx))
        if len(candidates) >= args.limit:
            break

    if not candidates:
        print("没有需要分析的工单（全部已在 Excel 中）。")
        return

    print(f"本次分析 {len(candidates)} 条（order={args.order}, refresh={args.refresh}）")
    client = AcpClient()
    analyzed_at = datetime.now(timezone.utc).isoformat()
    try:
        for i, (ticket, row_idx) in enumerate(candidates, start=1):
            tn = ticket["ticket_number"]
            print(f"[{i}/{len(candidates)}] ticket #{tn} ...", flush=True)
            row = {
                "ticket_number": tn,
                "title": ticket["title"],
                "status": ticket.get("status") or "",
                "urgency": ticket.get("urgency") or "",
                "description": (ticket.get("description") or "")[:3000],
                "issue_type": ticket.get("issue_type") or "",
                "business_line": ticket.get("business_line") or "",
                "ai_bug_category": ticket.get("ai_bug_category") or "",
                "record_id": ticket["record_id"],
                "source_updated_at": ticket.get("source_updated_at") or "",
                "analyzed_at": analyzed_at,
            }
            try:
                text = client.analyze(build_prompt(ticket))
                data = parse_result(text)
                intent = data["analysis"]["intent"]
                row.update({
                    "intent": intent["intentType"],
                    "sub_intent": intent.get("intentSubtype") or "",
                    "confidence": intent.get("confidence"),
                    "result_json": json.dumps(data, ensure_ascii=False),
                    "run_status": "ok",
                    "error": "",
                })
                print(f"    -> {row['intent']} / {row['sub_intent']} ({row['confidence']})")
            except Exception as exc:  # keep going on per-ticket failure
                row.update({"intent": "", "sub_intent": "", "confidence": "",
                            "result_json": "", "run_status": "error", "error": str(exc)[:500]})
                print(f"    -> ERROR: {exc}")
            values = [row.get(c, "") for c in COLUMNS]
            if row_idx is not None:
                for col, value in enumerate(values, start=1):
                    ws.cell(row=row_idx, column=col, value=value)
            else:
                ws.append(values)
            wb.save(excel_path)
    finally:
        client.close()

    print(f"已写入 {excel_path}")


if __name__ == "__main__":
    main()
